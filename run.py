import random
import time
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
import os

###Definindo as turmas e as matérias com suas respectivas cargas horárias
turmas= {
    "5A": {"Artes": 2, "Educação Física-Professor1": 3, "Inglês": 1, "Regente-Professor1": 7, "Regente-Professor2": 7},
    "5B": {"Artes": 2, "Educação Física-Professor2": 3, "Inglês": 1, "Regente-Professor1": 7, "Regente-Professor2": 7},
    "6A": {"Artes": 2, "Ciências-Professor1": 3, "Educação Física-Professor1": 2, "Ensino Religioso": 2, "Geografia": 1, "História": 2, "Inglês": 1, "Matemática-Professor1": 3, "Português-Professor1": 4},
    "6B": {"Artes": 2, "Ciências-Professor1": 3, "Educação Física-Professor1": 2, "Ensino Religioso": 2, "Geografia": 1, "História": 2, "Inglês": 1, "Matemática-Professor1": 3, "Português-Professor1": 4},
    "7A": {"Artes": 1, "Ciências-Professor1": 3, "Educação Física-Professor1": 2, "Ensino Religioso": 2, "Geografia": 2, "História": 1, "Inglês": 2, "Matemática-Professor1": 4, "Português-Professor1": 3},
    "7B": {"Artes": 1, "Ciências-Professor1": 3, "Educação Física-Professor1": 2, "Ensino Religioso": 2, "Geografia": 2, "História": 1, "Inglês": 2, "Matemática-Professor1": 4, "Português-Professor1": 3},
    "8A": {"Artes": 1, "Ciências-Professor2": 4, "Educação Física-Professor2": 2, "Ensino Religioso": 2, "Geografia": 2, "História": 2, "Inglês": 1, "Matemática-Professor2": 3, "Português-Professor2": 3},
    "8B": {"Artes": 1, "Ciências-Professor2": 4, "Educação Física-Professor2": 2, "Ensino Religioso": 2, "Geografia": 2, "História": 2, "Inglês": 1, "Matemática-Professor2": 3, "Português-Professor2": 3},
    "9A": {"Artes": 1, "Ciências-Professor2": 3, "Educação Física-Professor2": 2, "Ensino Religioso": 1, "Geografia": 2, "História": 2, "Inglês": 2, "Matemática-Professor2": 4, "Português-Professor2": 3},
    "9B": {"Artes": 1, "Ciências-Professor2": 3, "Educação Física-Professor2": 2, "Ensino Religioso": 1, "Geografia": 2, "História": 2, "Inglês": 2, "Matemática-Professor2": 4, "Português-Professor2": 3}
}

#Função para configurar e atualizar o gráfico
def configurar_grafico():
    plt.ion()  # Ativar modo interativo
    fig, ax = plt.subplots()
    ax.set_xlabel("Geração")
    ax.set_ylabel("Penalidade")
    ax.set_title("Penalidades x Geração")
    linha, = ax.plot([], [], 'b-', label="Melhor Penalidade")
    ax.legend()
    return fig, ax, linha

def atualizar_grafico(fig, ax, linha, geracoes, penalidades):
    linha.set_xdata(geracoes)
    linha.set_ydata(penalidades)
    ax.relim()  # Recalcular limites
    ax.autoscale_view()  # Ajustar escala automaticamente
    plt.draw()
    plt.pause(0.01)  # Pausar para atualizar o gráfico

# turmas= turmaExterno.balanceamento_professores_materia()
print(f"turmas:  {turmas}")

def gerar_horarios_todas_turmas_contador(num_individuos):
    for i in range(num_individuos):
        numero_individuo = num_individuos - i
        print(f"Indivíduo: {numero_individuo} criado.")
        yield gerar_horarios_todas_turmas()

# Gera os horários para todas as turmas
def gerar_horarios_todas_turmas():   
    def verificar_alocacao(horario, carga_horaria, turma):
        """Verifica se todas as matérias foram alocadas corretamente."""
        alocadas = [materia for dia in horario for materia in dia if materia]
        for materia, carga in carga_horaria.items():
            if alocadas.count(materia) != carga:
                print(f"Erro: A matéria {materia} não foi completamente alocada na turma {turma}.") 

    horarios_turmas = {}

    for turma, carga_horaria in turmas.items():
        horarios_turmas[turma] = gerar_horarios_turmas(carga_horaria)    
         
    verificar_alocacao(horarios_turmas[turma], carga_horaria, turma)
    
    return horarios_turmas

def introduzir_diversidade_com_mutacao(populacao, proporcao=0.3):
    num_mutacoes = int(len(populacao) * proporcao)
    for i in range(num_mutacoes):
        populacao[i] = (mutacao(populacao[i][0]), avaliar_aptidao(populacao[i][0]))
    return populacao

def salvar_horarios_em_excel(horarios):
    """
    Salva os horários gerados em um único arquivo Excel, com todas as turmas em uma aba.
    """
    # Mapeamento de cores para as matérias
    cores_materias = {
        "Artes": "C17DA5",  #  magenta suave
        "Educação Física-Professor1": "AFB87A",  # Verde Oliva
        "Educação Física-Professor2": "87CEEB",  # Azul céu
        "Inglês": "FFFFE0",  # Amarelo claro
        "Regente-Professor1": "98FB98",  # Verde claro
        "Regente-Professor2": "9CA4EC",  # azul claro com um toque de lavanda
        "Ciências-Professor1": "FFA07A",  # Salmão claro
        "Ciências-Professor2": "FA8072",  # Salmão
        "Ensino Religioso": "DDA0DD",  # Ameixa
        "Geografia": "FFD700",  # Ouro
        "História": "FFB6C1",  # Rosa claro
        "Matemática-Professor1": "00FA9A",  # Verde médio
        "Matemática-Professor2": "7CFC00",  # Verde grama
        "Português-Professor1": "FF6347",  # Tomate
        "Português-Professor2": "FF4500",  # Laranja vermelho
    }

    # Obter o diretório do script em execução
    if '__file__' in globals():
         current_dir = os.path.dirname(os.path.abspath(__file__))
    else:
        current_dir = os.getcwd()  
    file_name = os.path.join(current_dir, f"horarios_turmas_{time.time()}.xlsx")    

    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        dados_planilha = []

        for turma, dias in horarios.items():
            # Adicionar nome da turma como uma linha separadora
            dados_planilha.append([turma] + [""] * 5)

            # Adicionar cabeçalho dos dias da semana
            dados_planilha.append(["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Horário"])

            # Adicionar os horários
            for horario in range(4):  # Supondo 4 períodos por dia
                linha = [dias[dia_idx][horario] for dia_idx in range(5)] + [f"{horario + 1}º Horário"]
                dados_planilha.append(linha)

        # Criar DataFrame e salvar no Excel
        df_final = pd.DataFrame(dados_planilha)
        df_final.to_excel(writer, sheet_name="Horários", index=False, header=False)

    # Abrir o arquivo gerado para aplicar estilos
    wb = load_workbook(file_name)
    ws = wb.active

    # Estilo para colorir as linhas
    azul_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    bold_font = Font(bold=True)

        # Estilo de borda
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for i, row in enumerate(ws.iter_rows()):
        for cell in row:
            if cell.value:  # Aplicar borda apenas em células com valor
                cell.border = thin_border
        if row[0].value and row[0].value in horarios.keys():  # Linha com o nome da turma
            for cell in row:
                cell.fill = azul_fill
                cell.font = bold_font
            # Aplicar o mesmo estilo à próxima linha (dias da semana)
            for cell in ws[i + 2]:  # A linha dos dias da semana está logo após a linha da turma
                cell.fill = azul_fill
                cell.font = bold_font
        else:
            # Aplicar cores às matérias
            for cell in row:
                if cell.value in cores_materias:
                    cell.fill = PatternFill(start_color=cores_materias[cell.value], end_color=cores_materias[cell.value], fill_type="solid")

    # Salvar o arquivo com os estilos aplicados
    wb.save(file_name)
    print("Planilha Excel gerada com sucesso!")


def gerar_horarios_turmas(carga_horaria):    
    horario = [[None] * 4 for _ in range(5)]  # 5 dias, 4 períodos por dia       
    materias_com_peso = []    
    materias_adicionadas_turma = {materia: set() for materia in carga_horaria.keys()}  # Inicializa como dicionário de conjuntos
    horarios_ocupados = set()  # Conjunto para armazenar (dia, periodo) já ocupados
    horarios_disponiveis = [(dia, periodo) for dia in range(5) for periodo in range(4)]  # Todos os horários possíveis

    # Cria a lista de matérias com peso
    for materia, peso in carga_horaria.items():
        materias_com_peso.append({"materia": materia, "peso": peso})

    # Ordena as matérias com base no peso em ordem decrescente
    materias_com_peso.sort(key=lambda x: x["peso"], reverse=True)

    # Distribui as matérias
    for materia_info in materias_com_peso:
        materia = materia_info["materia"]
        peso = materia_info["peso"]
        for _ in range(peso):
            alocado = False
            # Atualiza os horários livres removendo os ocupados
            horarios_livres = [h for h in horarios_disponiveis if h not in horarios_ocupados]

            while not alocado:
                if not horarios_livres:
                    raise ValueError("Não há horários livres suficientes para alocar todas as matérias.")
                
                # Escolhe um horário aleatório dos horários livres
                dia, periodo = random.choice(horarios_livres)
                # carga_horaria_semanal_materia = len(materias_adicionadas_turma[materia])

                # Verifica se o horário está vazio e se a matéria pode ser alocada nesse dia
                if materia in ["Regente-Professor1", "Regente-Professor2"] or dia not in materias_adicionadas_turma[materia]:                 
                    horario[dia][periodo] = materia
                    materias_adicionadas_turma[materia].add(dia)                    
                    horarios_ocupados.add((dia, periodo))  # Marca o horário como ocupado
                    alocado = True                    
                else:
                    dias_alocados = materias_adicionadas_turma.get(materia, set()) 
                    dias_disponiveis = {dia for dia, _ in horarios_livres}
                    
                    # Verifica os dias disponíveis que não estão nos dias já alocados
                    dias_diferentes = dias_disponiveis - dias_alocados
                    
                    if not dias_diferentes:                                                           
                        # Escolhe uma matéria aleatória já alocada para desalocar
                        materia_para_desalocar = random.choice(list(materias_adicionadas_turma.keys()))
                        if materias_adicionadas_turma[materia_para_desalocar]:
                            dia_para_remover = materias_adicionadas_turma[materia_para_desalocar].pop()
                            
                            # Encontra o período correspondente e remove a matéria do horário
                            for periodo_idx in range(4):
                                if horario[dia_para_remover][periodo_idx] == materia_para_desalocar:
                                    horario[dia_para_remover][periodo_idx] = None
                                    horarios_ocupados.remove((dia_para_remover, periodo_idx))
                                    break
                            
                            # Adiciona o peso de volta para a matéria desalocada
                            for materia_info in materias_com_peso:
                                if materia_info["materia"] == materia_para_desalocar:
                                    materia_info["peso"] += 1
                                    break
                        else:
                            alocado = False

    return horario

def avaliar_aptidao(horarios):
    penalidades = 0
       
    # Dicionário para rastrear a carga horária de cada matéria em todas as turmas
    carga_horaria_geral = {}


    for turma, horario in horarios.items():
        # 1- Carga horária da matéria sendo cumprida por turmas
        carga_horaria_gerada = {}
        for dia in horario:
            for materia in dia:
                if materia:
                    carga_horaria_gerada[materia] = carga_horaria_gerada.get(materia, 0) + 1
        for materia, carga in turmas[turma].items():
            if carga_horaria_gerada.get(materia, 0) != carga:
                # penalidades += abs(carga_horaria_gerada.get(materia, 0) - carga) * 5  # Penalidade maior por erro de carga horária
                penalidades += 4  # Penalidade maior por erro de carga horária
        
        # 2- Verificar repetição de matérias no mesmo dia
        for dia_idx, dia in enumerate(horario):
            materias_no_dia = set()
            carga_horaria_dia = {}  # Contagem de aulas por matéria no dia          
            for periodo_idx, materia in enumerate(dia):              
                if materia:
                    # 3- Verificar repetição de matérias no mesmo dia
                    # Penalizar repetição de matérias no mesmo dia
                    if turma == "5A" and materia not in ["Regente-Professor1", "Regente-Professor2"] and materia in materias_no_dia:
                        penalidades += 1  # Penalidade para repetição de matérias no 5A (exceto regentes)
                    elif turma == "5B" and materia not in ["Regente-Professor1", "Regente-Professor2"] and materia in materias_no_dia:
                        penalidades += 1  # Penalidade para repetição de matérias no 5B (exceto regentes)
                    elif turma not in ["5A", "5B"] and materia in materias_no_dia:
                        penalidades += 1  # Penalidade para repetição de qualquer matéria nas outras turmas
                    materias_no_dia.add(materia)


    # 3- Após processar todas as turmas, verificar excesso de carga horária por matéria no dia
    for turma, horario in horarios.items():
        for dia_idx, dia in enumerate(horario):
            for periodo_idx, materia in enumerate(dia):
                if materia:
                    # Atualizar a carga horária geral da matéria
                    if materia not in carga_horaria_geral:
                        carga_horaria_geral[materia] = {}
                    if dia_idx not in carga_horaria_geral[materia]:
                        carga_horaria_geral[materia][dia_idx] = 0
                    carga_horaria_geral[materia][dia_idx] += 1

    # 3.a- Verificar excesso de carga horária por matéria no dia
    for materia, dias in carga_horaria_geral.items():
        for dia, total_periodos in dias.items():
            if total_periodos > 4:  # Limite de 4 períodos por dia
                penalidades += (total_periodos - 4) * 5  # Penalidade para excesso de carga horária

    # 4- Verificar se a mesma matéria está alocada no mesmo horário em turmas diferentes
    horarios_por_periodo = {}  # Dicionário para mapear (dia, período) para matérias

    for turma, horario in horarios.items():
        for dia_idx, dia in enumerate(horario):
            for periodo_idx, materia in enumerate(dia):
                if materia:
                    if (dia_idx, periodo_idx) not in horarios_por_periodo:
                        horarios_por_periodo[(dia_idx, periodo_idx)] = set()
                    if materia in horarios_por_periodo[(dia_idx, periodo_idx)]:
                        penalidades += 3  # Penalidade por conflito de horários
                    horarios_por_periodo[(dia_idx, periodo_idx)].add(materia)
    
    return penalidades

# Seleção por ranking
def selecao_por_ranking(populacao, num_individuos):
    # Ordenar a população pela aptidão (menor penalidade primeiro)
    populacao_ordenada = sorted(populacao, key=lambda x: x[1])
    
    # Criar uma lista de probabilidades inversamente proporcionais à posição no ranking
    ranking = list(range(1, len(populacao_ordenada) + 1))
    probabilidades = [1 / rank for rank in ranking]
    soma_probabilidades = sum(probabilidades)
    probabilidades_normalizadas = [p / soma_probabilidades for p in probabilidades]
    
    # Selecionar indivíduos com base nas probabilidades normalizadas
    selecionados = random.choices(
        populacao_ordenada, 
        weights=probabilidades_normalizadas, 
        k=num_individuos
    )
    
    # Retornar apenas os indivíduos selecionados (sem as penalidades)
    return [individuo[0] for individuo in selecionados]

# # # Cruzamento de dois pontos
def cruzamento(pai1, pai2):
    filho = {}
    for turma in pai1:
        horario_filho = []
        for dia1, dia2 in zip(pai1[turma], pai2[turma]):
            horario_filho.append([random.choice([a, b]) for a, b in zip(dia1, dia2)])
        filho[turma] = horario_filho
    return filho

# Aumentando a probabilidade de mutação
def mutacao(horarios):
    turma_mutada = random.choice(list(horarios.keys()))
    dia1, dia2 = random.sample(range(5), 2)  # Escolhe dois dias aleatórios
    periodo1, periodo2 = random.sample(range(4), 2)  # Escolhe dois períodos aleatórios
    # Troca as matérias entre os dois períodos
    horarios[turma_mutada][dia1][periodo1], horarios[turma_mutada][dia2][periodo2] = \
        horarios[turma_mutada][dia2][periodo2], horarios[turma_mutada][dia1][periodo1]
    return horarios

# Algoritmo Genético para gerar o horário
def algoritmo_genetico():
    # Parâmetros do algoritmo genético
    num_geracoes = 1300
    num_individuos = 800
    taxa_elitismo = 0.3
    proporcao_divsersidade_mudacao=0.3
    
	# Configurar gráfico
    fig, ax, linha = configurar_grafico()
    geracoes = []
    penalidades = []

    # Geração de população inicial
    print("começou gerar horario turmas")
    populacao = [(horario, 0) for horario in gerar_horarios_todas_turmas_contador(num_individuos)]
    print("finalizou gerar horario turmas")

    # Avaliar aptidão de cada indivíduo na população
    populacao = [(horarios, avaliar_aptidao(horarios)) for horarios, _ in populacao]

    # Inicializar a variável de melhor penalidade
    melhor_penalidade_anterior = float('inf')
    estagnacao = 0  # Inicializar antes do loop

    for geracao in range(num_geracoes):
        # # # Ajustar a taxa de mutação dinamicamente
        if estagnacao >= 6:
            probabilidade_mutacao = min(1.0, probabilidade_mutacao + 0.2)
        else:
            probabilidade_mutacao = 0.2

        # Seleção dos pais
        # pais = selecao_por_torneio(populacao, num_individuos, tamanho_torneio=5)
        pais = selecao_por_ranking(populacao, num_individuos)

        # Cruzamento dos pais para gerar filhos
        filhos = []
        for i in range(0, len(pais) - 1, 2):
            filho = cruzamento(pais[i], pais[i + 1])
            filhos.append(filho)

        if len(pais) % 2 != 0:
            filhos.append(pais[-1])

        # Mutação dos filhos
        for filho in filhos:
            if random.random() < probabilidade_mutacao:
                mutacao(filho)

        # Avaliação de aptidão dos filhos
        filhos = [(filho, avaliar_aptidao(filho)) for filho in filhos]

        # # Preservar a elite
        elite = populacao[:int(len(populacao) * taxa_elitismo)]
        populacao = elite + filhos

        ###Introduzir diversidade se necessário
        if estagnacao >= 15:
            print("Reiniciando parte da população com maior diversidade...")
            populacao = introduzir_diversidade_com_mutacao(populacao, proporcao_divsersidade_mudacao)
            estagnacao = 0

        # Melhor indivíduo da geração
        melhor_individuo = min(populacao, key=lambda x: x[1])
        print(f"Geração {geracao + 1}: Penalidade = {melhor_individuo[1]} Probabilidade de mutação: {probabilidade_mutacao} ")

        #Atualizar Gráfico
        geracoes.append(geracao + 1)
        penalidades.append(melhor_individuo[1])
        atualizar_grafico(fig, ax, linha, geracoes, penalidades)
		
        # Verificar estagnação
        if melhor_individuo[1] < melhor_penalidade_anterior:
            melhor_penalidade_anterior = melhor_individuo[1]
            estagnacao = 0
        else:
            estagnacao += 1

        if melhor_individuo[1] == 0:
            print(f"Melhor horário encontrado na geração {geracao + 1}:")
            salvar_horarios_em_excel(melhor_individuo[0])
            break

def format_time(timestamp):
    """Formata um timestamp no formato HH:MM:SS."""
    return time.strftime("%H:%M:%S", time.localtime(timestamp))

def print_time(label, timestamp):
    """Imprime um timestamp formatado com um rótulo."""
    formatted_time = format_time(timestamp)
    print(f"{label}: {formatted_time}")

def run_algorithm(algorithm):
    """Executa um algoritmo e imprime o tempo inicial, final e o tempo total de execução."""
    start_time = time.time()
    print_time("Tempo inicial", start_time)
    
    algorithm()  # Executa o algoritmo passado como parâmetro
    
    end_time = time.time()
    print_time("Tempo final", end_time)
    
    # Calcula e imprime o tempo total de execução
    execution_time = end_time - start_time
    print(f"Tempo total de execução: {execution_time:.2f} segundos")

def main():
    print("Iniciando execução do algoritmo genético.")
    run_algorithm(algoritmo_genetico)
    print("Execução concluída.")

if __name__ == "__main__":
    main()
    plt.ioff()     
    plt.show()
